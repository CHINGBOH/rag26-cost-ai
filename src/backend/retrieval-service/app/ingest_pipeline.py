"""Issue #69 Phase A: ingest pipeline backed by PG state.

Extractors output a uniform Block contract:
    {doc_id, page, block_id, type, text, bbox?, metadata}

State machine:
    queued -> extracting -> chunking -> embedding -> indexing -> done
                                                              \-> failed

Topology principle: workers are stateless; PG ingest_jobs + ingest_write_log
are the source of truth. Killing/restarting the process must not cause
duplicate writes (write_log PK guards each phase write).
"""
from __future__ import annotations

import json
import logging
import os
import time
import uuid
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Iterable

logger = logging.getLogger("ingest_pipeline")


# ────────────────────────────── Block contract ──────────────────────────────
@dataclass
class Block:
    doc_id: str
    page: int
    block_id: str
    type: str           # text | table | figure | caption
    text: str
    bbox: tuple | None = None
    metadata: dict = field(default_factory=dict)


# ────────────────────────────── PG helpers ──────────────────────────────
def _conn():
    from app.agent.tools import _get_pg_conn
    return _get_pg_conn()


def _release(conn):
    from app.agent.tools import _put_pg_conn
    _put_pg_conn(conn)


# ────────────────────────────── Qdrant + Neo4j helpers ──────────────────────────────
QDRANT_URL = os.getenv("QDRANT_URL", "http://localhost:6333")
QDRANT_COLLECTION = os.getenv("QDRANT_INGEST_COLLECTION", "document_chunks")
NEO4J_URI = os.getenv("NEO4J_URI", "bolt://localhost:7687")
NEO4J_USER = os.getenv("NEO4J_USER", "neo4j")
NEO4J_PASS = os.getenv("NEO4J_PASSWORD", "password")


def _qdrant_point_id(doc_id: str, chunk_index: int) -> str:
    return str(uuid.uuid5(uuid.NAMESPACE_URL, f"{doc_id}#{chunk_index}"))


def qdrant_upsert(doc_id: str, file_name: str, chunk_index: int,
                  text: str, vec: list[float], page: int, block_type: str,
                  job_id: str) -> bool:
    """Upsert one point. Stable id ⇒ idempotent."""
    import httpx
    pid = _qdrant_point_id(doc_id, chunk_index)
    body = {"points": [{
        "id": pid,
        "vector": vec,
        "payload": {
            "doc_id": doc_id, "file_name": file_name,
            "chunk_index": chunk_index, "page": page,
            "block_type": block_type, "text": text[:500],
            "job_id": job_id,
        },
    }]}
    try:
        with httpx.Client(timeout=10.0, transport=httpx.HTTPTransport(proxy=None)) as cli:
            r = cli.put(f"{QDRANT_URL}/collections/{QDRANT_COLLECTION}/points?wait=true",
                        json=body)
            r.raise_for_status()
        return True
    except Exception as e:
        logger.warning("qdrant upsert failed (%s#%s): %s", doc_id, chunk_index, e)
        return False


def _neo4j_session():
    try:
        from neo4j import GraphDatabase  # type: ignore
    except Exception as e:
        logger.warning("neo4j driver unavailable: %s", e)
        return None
    try:
        drv = GraphDatabase.driver(NEO4J_URI, auth=(NEO4J_USER, NEO4J_PASS))
        drv.verify_connectivity()
        return drv
    except Exception as e:
        logger.warning("neo4j connect failed: %s", e)
        return None


# Lightweight regex-based entity extractor.
_MATERIAL_DICT = [
    "中砂", "粗砂", "细砂", "水泥", "钢筋", "混凝土", "砂浆", "电缆", "电线",
    "脚手架", "模板", "门窗", "管道", "保温", "防水", "瓷砖", "腻子", "涂料",
]


def _extract_entities(text: str) -> list[dict]:
    """Return [{type, name}] entities found via regex/dict."""
    import re
    ents: list[dict] = []
    seen: set[tuple[str, str]] = set()
    for m in _MATERIAL_DICT:
        if m in text:
            k = ("Material", m)
            if k not in seen:
                seen.add(k)
                ents.append({"type": "Material", "name": m})
    # Year-month
    for ym in re.findall(r"(20\d{2})[-年./]?\s*(\d{1,2})\s*月?", text):
        name = f"{ym[0]}-{int(ym[1]):02d}"
        k = ("Period", name)
        if k not in seen:
            seen.add(k)
            ents.append({"type": "Period", "name": name})
    # Section ids like 1.1, 2.3.4
    for sec in re.findall(r"(?<![\d.])\d+(?:\.\d+){1,3}(?![\d.])", text):
        k = ("Section", sec)
        if k not in seen:
            seen.add(k)
            ents.append({"type": "Section", "name": sec})
    return ents[:20]


def neo4j_write_chunk(driver, doc_id: str, file_name: str,
                      chunk_index: int, text: str, page: int,
                      block_type: str) -> int:
    """MERGE Document, Chunk, and entity relationships. Returns triples written."""
    if driver is None:
        return 0
    ents = _extract_entities(text)
    chunk_uid = f"{doc_id}#{chunk_index}"
    cypher = """
    MERGE (d:Document {doc_id: $doc_id})
      ON CREATE SET d.file_name = $file_name
      ON MATCH  SET d.file_name = $file_name
    MERGE (c:Chunk {uid: $cuid})
      ON CREATE SET c.doc_id = $doc_id, c.chunk_index = $cidx, c.page = $page,
                    c.block_type = $btype, c.text = left($text, 200)
      ON MATCH  SET c.page = $page, c.block_type = $btype, c.text = left($text, 200)
    MERGE (d)-[:HAS_CHUNK]->(c)
    WITH c
    UNWIND $ents AS e
      CALL {
        WITH c, e
        MERGE (n:Entity {type: e.type, name: e.name})
        MERGE (c)-[:MENTIONS]->(n)
        RETURN count(*) AS x
      }
    RETURN sum(x) AS rels
    """
    try:
        with driver.session() as ses:
            res = ses.run(cypher, doc_id=doc_id, file_name=file_name,
                          cuid=chunk_uid, cidx=chunk_index, page=page,
                          btype=block_type, text=text, ents=ents)
            rec = res.single()
            return (rec["rels"] if rec and rec.get("rels") else 0) + 2  # 2 = HAS_CHUNK + Chunk
    except Exception as e:
        logger.warning("neo4j write failed (%s#%s): %s", doc_id, chunk_index, e)
        return 0


def job_create(file_name: str, file_path: str, file_size: int, mime: str) -> str:
    job_id = uuid.uuid4().hex[:16]
    conn = _conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO ingest_jobs
                  (job_id, doc_id, file_name, file_path, file_size, mime, status, phase)
                VALUES (%s, %s, %s, %s, %s, %s, 'queued', 'queued')
                """,
                (job_id, job_id, file_name, file_path, file_size, mime),
            )
        conn.commit()
    finally:
        _release(conn)
    return job_id


def job_update(job_id: str, **fields) -> None:
    if not fields:
        return
    cols = list(fields.keys())
    vals = list(fields.values())
    sets = ", ".join(f"{c} = %s" for c in cols)
    conn = _conn()
    try:
        with conn.cursor() as cur:
            cur.execute(f"UPDATE ingest_jobs SET {sets} WHERE job_id = %s", vals + [job_id])
        conn.commit()
    finally:
        _release(conn)


def job_get(job_id: str) -> dict | None:
    conn = _conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT job_id, doc_id, file_name, file_path, file_size, mime, status,
                          phase, progress_pct, error, blocks_total, chunks_pg,
                          vectors_qdrant, triples_neo4j, ocr_pages, text_chars,
                          extractor, duration_ms, created_at, updated_at,
                          started_at, finished_at, metadata
                   FROM ingest_jobs WHERE job_id = %s""",
                (job_id,),
            )
            row = cur.fetchone()
            if not row:
                return None
            cols = [d[0] for d in cur.description]
            return dict(zip(cols, row))
    finally:
        _release(conn)


def job_list(limit: int = 50, status: str | None = None) -> list[dict]:
    conn = _conn()
    try:
        with conn.cursor() as cur:
            if status:
                cur.execute(
                    """SELECT job_id, doc_id, file_name, file_size, mime, status, phase,
                              progress_pct, blocks_total, chunks_pg, vectors_qdrant,
                              extractor, duration_ms, created_at, updated_at, error
                       FROM ingest_jobs WHERE status = %s
                       ORDER BY created_at DESC LIMIT %s""",
                    (status, limit),
                )
            else:
                cur.execute(
                    """SELECT job_id, doc_id, file_name, file_size, mime, status, phase,
                              progress_pct, blocks_total, chunks_pg, vectors_qdrant,
                              extractor, duration_ms, created_at, updated_at, error
                       FROM ingest_jobs
                       ORDER BY created_at DESC LIMIT %s""",
                    (limit,),
                )
            cols = [d[0] for d in cur.description]
            return [dict(zip(cols, r)) for r in cur.fetchall()]
    finally:
        _release(conn)


def write_log_has(job_id: str, phase: str, key: str) -> bool:
    conn = _conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT 1 FROM ingest_write_log WHERE job_id=%s AND phase=%s AND key=%s",
                (job_id, phase, key),
            )
            return cur.fetchone() is not None
    finally:
        _release(conn)


def write_log_put(job_id: str, phase: str, key: str) -> None:
    conn = _conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """INSERT INTO ingest_write_log (job_id, phase, key)
                   VALUES (%s, %s, %s) ON CONFLICT DO NOTHING""",
                (job_id, phase, key),
            )
        conn.commit()
    finally:
        _release(conn)


def blindspots_for_job(job_id: str) -> list[dict]:
    conn = _conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT page, kind, reason, image_count, text_chars
                     FROM ingest_blindspots
                    WHERE job_id = %s
                 ORDER BY page""",
                (job_id,),
            )
            return [
                {"page": r[0], "kind": r[1], "reason": r[2],
                 "image_count": r[3], "text_chars": r[4]}
                for r in cur.fetchall()
            ]
    except Exception as e:
        logger.warning("blindspots_for_job failed: %s", e)
        return []
    finally:
        _release(conn)


def blindspots_list(doc_id: str | None = None, limit: int = 100) -> list[dict]:
    conn = _conn()
    try:
        with conn.cursor() as cur:
            if doc_id:
                cur.execute(
                    """SELECT job_id, doc_id, file_name, page, kind, reason,
                              image_count, text_chars, detected_at
                         FROM ingest_blindspots
                        WHERE doc_id = %s
                     ORDER BY page LIMIT %s""",
                    (doc_id, limit),
                )
            else:
                cur.execute(
                    """SELECT job_id, doc_id, file_name, page, kind, reason,
                              image_count, text_chars, detected_at
                         FROM ingest_blindspots
                     ORDER BY detected_at DESC LIMIT %s""",
                    (limit,),
                )
            return [
                {"job_id": r[0], "doc_id": r[1], "file_name": r[2],
                 "page": r[3], "kind": r[4], "reason": r[5],
                 "image_count": r[6], "text_chars": r[7],
                 "detected_at": str(r[8]) if r[8] else None}
                for r in cur.fetchall()
            ]
    except Exception as e:
        logger.warning("blindspots_list failed: %s", e)
        return []
    finally:
        _release(conn)


# ────────────────────────────── extractors ──────────────────────────────
def extract_text_file(path: Path, doc_id: str) -> tuple[list[Block], str]:
    """Plain .txt / .md — one block per non-empty paragraph."""
    text = path.read_text(encoding="utf-8", errors="ignore")
    blocks: list[Block] = []
    for i, para in enumerate(p.strip() for p in text.split("\n\n")):
        if para:
            blocks.append(Block(doc_id=doc_id, page=1, block_id=f"p1_b{i}",
                                type="text", text=para))
    return blocks, "text"


def extract_csv(path: Path, doc_id: str) -> tuple[list[Block], str]:
    """CSV — one Block per row, plus a header summary block."""
    import csv
    blocks: list[Block] = []
    rows: list[list[str]] = []
    with open(path, newline="", encoding="utf-8", errors="ignore") as fh:
        # try utf-8-sig in case of BOM
        try:
            reader = csv.reader(fh)
            for row in reader:
                rows.append([(c or "").strip() for c in row])
        except Exception as e:
            raise RuntimeError(f"csv parse failed: {e}")

    if not rows:
        return [], "csv"

    header = rows[0]
    blocks.append(Block(
        doc_id=doc_id, page=1, block_id="hdr",
        type="caption", text="表头: " + " | ".join(header),
        metadata={"row": 0, "header": header},
    ))
    for ri, row in enumerate(rows[1:], start=1):
        if not any(row):
            continue
        # render as "key: val | key: val" for embedding-friendliness
        cells = [f"{header[i] if i < len(header) else f'col{i}'}: {v}"
                 for i, v in enumerate(row) if v]
        blocks.append(Block(
            doc_id=doc_id, page=1, block_id=f"row{ri}",
            type="table_row", text=" | ".join(cells),
            metadata={"row": ri},
        ))
    return blocks, "csv"


def extract_xlsx(path: Path, doc_id: str) -> tuple[list[Block], str]:
    """Excel — each sheet → caption block + one table_row block per data row."""
    import openpyxl
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    blocks: list[Block] = []
    page = 0
    try:
        for sname in wb.sheetnames:
            page += 1
            ws = wb[sname]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_raw = next(rows_iter)
            except StopIteration:
                continue
            header = [str(c) if c is not None else "" for c in header_raw]
            blocks.append(Block(
                doc_id=doc_id, page=page, block_id=f"sheet{page}_hdr",
                type="caption", text=f"工作表 [{sname}] 表头: " + " | ".join(h for h in header if h),
                metadata={"sheet": sname, "row": 0, "header": header},
            ))
            for ri, row in enumerate(rows_iter, start=1):
                cells_raw = [("" if v is None else str(v)).strip() for v in row]
                if not any(cells_raw):
                    continue
                cells = [f"{header[i] if i < len(header) else f'col{i}'}: {v}"
                         for i, v in enumerate(cells_raw) if v]
                if not cells:
                    continue
                blocks.append(Block(
                    doc_id=doc_id, page=page, block_id=f"sheet{page}_r{ri}",
                    type="table_row", text=" | ".join(cells),
                    metadata={"sheet": sname, "row": ri},
                ))
    finally:
        wb.close()
    return blocks, "xlsx"


def _extract_pdf_tables(path: Path, doc_id: str) -> list[Block]:
    """pdfplumber-based table extraction; merges per page, robust to failures."""
    blocks: list[Block] = []
    try:
        import pdfplumber
    except Exception:
        return blocks
    try:
        with pdfplumber.open(str(path)) as pdf:
            for pi, page in enumerate(pdf.pages, start=1):
                try:
                    tables = page.extract_tables() or []
                except Exception:
                    continue
                for ti, tbl in enumerate(tables):
                    if not tbl or not any(any(c for c in r) for r in tbl):
                        continue
                    header = [str(c or "").strip() for c in tbl[0]]
                    blocks.append(Block(
                        doc_id=doc_id, page=pi,
                        block_id=f"p{pi}_t{ti}_hdr",
                        type="caption",
                        text=f"表 {pi}-{ti+1} 表头: " + " | ".join(h for h in header if h),
                        metadata={"table": ti, "row": 0, "header": header},
                    ))
                    for ri, row in enumerate(tbl[1:], start=1):
                        cells_raw = [str(c or "").strip() for c in row]
                        if not any(cells_raw):
                            continue
                        cells = [f"{header[i] if i < len(header) else f'col{i}'}: {v}"
                                 for i, v in enumerate(cells_raw) if v]
                        if not cells:
                            continue
                        blocks.append(Block(
                            doc_id=doc_id, page=pi,
                            block_id=f"p{pi}_t{ti}_r{ri}",
                            type="table_row",
                            text=" | ".join(cells),
                            metadata={"table": ti, "row": ri},
                        ))
    except Exception as e:
        logger.warning("pdfplumber table extraction failed: %s", e)
    return blocks


def record_blindspot(job_id: str, doc_id: str, file_name: str, page: int,
                     kind: str, reason: str = "", bbox: dict | None = None,
                     image_count: int = 0, text_chars: int = 0) -> None:
    """UPSERT a blindspot row. Idempotent via UNIQUE(doc_id, page, kind)."""
    conn = _conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """INSERT INTO ingest_blindspots
                     (job_id, doc_id, file_name, page, kind, reason, bbox,
                      image_count, text_chars)
                   VALUES (%s, %s, %s, %s, %s, %s, %s::jsonb, %s, %s)
                   ON CONFLICT (doc_id, page, kind) DO UPDATE
                     SET reason = EXCLUDED.reason,
                         image_count = EXCLUDED.image_count,
                         text_chars = EXCLUDED.text_chars""",
                (job_id, doc_id, file_name, page, kind, reason,
                 json.dumps(bbox) if bbox else None,
                 image_count, text_chars),
            )
        conn.commit()
    except Exception as e:
        logger.warning("blindspot insert failed: %s", e)
    finally:
        _release(conn)


def extract_pdf_native(path: Path, doc_id: str,
                       job_id: str | None = None,
                       file_name: str = "",
                       force_scan_fallback: bool = False) -> tuple[list[Block], str] | None:
    """Try PyMuPDF first; if pages have very little text, signal caller to OCR.

    Streams page-by-page so 905M files don't blow up memory. Pulls native
    tables via pdfplumber and merges into the block stream. Pages with images
    but very little text are recorded as blindspots.

    When ``force_scan_fallback`` is True, returns whatever sparse text we got
    plus a placeholder block per empty page (so chunker has something) and
    flags every page as a scan_no_ocr blindspot. Used when OCR service is
    unreachable so the job finishes degraded rather than failing the topology.
    """
    try:
        import fitz  # PyMuPDF
    except Exception:
        return None
    blocks: list[Block] = []
    total_text_chars = 0
    n_pages = 0
    blindspots: list[dict] = []
    try:
        with fitz.open(str(path)) as doc:
            n_pages = doc.page_count
            for pi in range(n_pages):
                page = doc.load_page(pi)
                txt = page.get_text("text") or ""
                total_text_chars += len(txt)
                # Detect images on this page — chart/figure heuristic
                try:
                    n_img = len(page.get_images(full=False) or [])
                except Exception:
                    n_img = 0
                page_chars = len(txt.strip())
                # Heuristic: image present + sparse text → likely chart/figure
                if n_img >= 1 and page_chars < 80:
                    blindspots.append({
                        "page": pi + 1, "kind": "chart" if n_img == 1 else "figure",
                        "reason": f"page has {n_img} image(s) but only {page_chars} text chars",
                        "image_count": n_img, "text_chars": page_chars,
                    })
                for bi, para in enumerate(p.strip() for p in txt.split("\n\n")):
                    if para and len(para) > 4:
                        blocks.append(Block(
                            doc_id=doc_id, page=pi + 1,
                            block_id=f"p{pi+1}_b{bi}",
                            type="text", text=para,
                        ))
    except Exception as e:
        logger.warning("pymupdf failed: %s", e)
        return None
    if not blocks or (total_text_chars / max(1, n_pages)) < 30:
        if not force_scan_fallback:
            return None  # signal scan / needs OCR
        # Degraded path: surface whatever sparse text we got + placeholders
        # for empty pages, and flag every page as scan_no_ocr.
        logger.warning("native fallback: %d pages, %d total chars (avg %.1f) — degraded mode",
                       n_pages, total_text_chars, total_text_chars / max(1, n_pages))
        seen_pages = {b.page for b in blocks}
        for pi in range(1, n_pages + 1):
            if pi not in seen_pages:
                blocks.append(Block(
                    doc_id=doc_id, page=pi,
                    block_id=f"p{pi}_scan_placeholder",
                    type="caption",
                    text=f"[第{pi}页：扫描图未OCR，原始文本为空]",
                    metadata={"scan_no_ocr": True},
                ))
        if job_id:
            existing_pages = {bs["page"] for bs in blindspots}
            for pi in range(1, n_pages + 1):
                if pi in existing_pages:
                    continue
                blindspots.append({
                    "page": pi, "kind": "scan_no_ocr",
                    "reason": "OCR service unreachable; native PDF text is sparse",
                    "image_count": 0, "text_chars": 0,
                })

    # native PDF: also try to pull tables (best-effort, never fatal)
    try:
        tbl_blocks = _extract_pdf_tables(path, doc_id)
        if tbl_blocks:
            blocks.extend(tbl_blocks)
            blocks.sort(key=lambda b: (b.page, 0 if b.type == "text" else 1))
            logger.info("pdfplumber added %d table blocks", len(tbl_blocks))
    except Exception as e:
        logger.warning("table extraction skipped: %s", e)

    # Persist blindspots (best-effort)
    if job_id and blindspots:
        for bs in blindspots:
            record_blindspot(
                job_id=job_id, doc_id=doc_id, file_name=file_name,
                page=bs["page"], kind=bs["kind"], reason=bs["reason"],
                image_count=bs["image_count"], text_chars=bs["text_chars"],
            )
        logger.info("recorded %d blindspots for doc %s", len(blindspots), doc_id)

    return blocks, "pymupdf"


async def extract_pdf_or_image_via_ocr(path: Path, doc_id: str, file_name: str,
                                       job_id: str | None = None) -> tuple[list[Block], str]:
    """Call OCR service, then convert PaddleOCR result -> IngestDocument
    -> legacy Block stream. Consumes text_blocks / tables / figures so
    PPStructure output is no longer thrown away.

    On unreachable OCR / non-2xx response, degrades to extract_pdf_native
    with force_scan_fallback=True so the pipeline never hard-fails.
    """
    import httpx
    from app.ingest_schema import from_paddle_ocr_json, to_legacy_blocks

    ocr_url = os.environ.get("OCR_SERVICE_URL", "http://localhost:8001")
    suffix = path.suffix.lower()
    endpoint = "/ocr/pdf" if suffix == ".pdf" else "/ocr/image"

    try:
        async with httpx.AsyncClient(timeout=600.0,
                                     transport=httpx.AsyncHTTPTransport(proxy=None)) as client:
            with open(path, "rb") as fh:
                r = await client.post(
                    f"{ocr_url}{endpoint}",
                    files={"file": (file_name, fh)},
                )
        if r.status_code != 200:
            raise RuntimeError(f"OCR HTTP {r.status_code}: {r.text[:200]}")
        data = r.json()
    except Exception as e:
        logger.warning("OCR unreachable / failed (%s); falling back to native + scan_no_ocr blindspots",
                       type(e).__name__)
        if suffix == ".pdf":
            res = extract_pdf_native(path, doc_id, job_id=job_id, file_name=file_name,
                                     force_scan_fallback=True)
            if res:
                blocks, _extractor = res
                return blocks, "scan_no_ocr_fallback"
        # Image with no OCR: emit one caption block + a blindspot so the
        # job finishes degraded instead of crashing.
        if job_id:
            try:
                record_blindspot(
                    job_id=job_id, doc_id=doc_id, file_name=file_name,
                    page=1, kind="scan_no_ocr",
                    reason="OCR service unreachable; image cannot be parsed",
                    image_count=1, text_chars=0,
                )
            except Exception:
                pass
        return [Block(
            doc_id=doc_id, page=1, block_id="p1_scan_placeholder",
            type="caption",
            text=f"[图像 {file_name}：OCR 服务不可达，未识别文字]",
            metadata={"scan_no_ocr": True},
        )], "scan_no_ocr_fallback"

    # Convert via schema → blocks
    doc = from_paddle_ocr_json(data, doc_id=doc_id, file_name=file_name)
    blocks = to_legacy_blocks(doc)

    # Persist blindspots for any page that has only figure/caption blocks
    # (pure chart pages) so the agent can disclose data gaps.
    if job_id:
        for page in doc.pages:
            text_blocks = [b for b in page.blocks if b.type == "text" and len(b.text.strip()) >= 20]
            fig_blocks = [b for b in page.blocks if b.type == "figure"]
            if fig_blocks and not text_blocks:
                try:
                    record_blindspot(
                        job_id=job_id, doc_id=doc_id, file_name=file_name,
                        page=page.page,
                        kind="chart" if len(fig_blocks) == 1 else "figure",
                        reason=f"OCR found {len(fig_blocks)} figure(s) but no text on page",
                        image_count=len(fig_blocks),
                        text_chars=sum(len((b.text or "").strip()) for b in page.blocks),
                    )
                except Exception as e:
                    logger.debug("blindspot record skipped: %s", e)

    logger.info("OCR adapter: %d pages -> %d blocks (text/table/figure split)",
                len(doc.pages), len(blocks))
    return blocks, "ocr_paddle"


# ────────────────────────────── chunking ──────────────────────────────
def chunk_blocks(blocks: list[Block], max_chars: int = 400) -> list[Block]:
    """Page-aware chunker.

    - text blocks: merged within page until max_chars.
    - table_row / caption / figure: kept as their own chunk so retrieval
      can pinpoint a row without diluting the embedding.
    """
    out: list[Block] = []
    cur_buf: list[str] = []
    cur_page = -1
    cur_doc_id = ""

    def flush():
        nonlocal cur_buf, cur_page, cur_doc_id
        if cur_buf:
            txt = "\n\n".join(cur_buf)
            out.append(Block(
                doc_id=cur_doc_id, page=cur_page,
                block_id=f"chunk_{len(out)}",
                type="text", text=txt,
            ))
            cur_buf = []

    for b in blocks:
        cur_doc_id = b.doc_id
        # non-text blocks: emit as-is, bypass buffer
        if b.type != "text":
            flush()
            out.append(Block(
                doc_id=b.doc_id, page=b.page,
                block_id=f"chunk_{len(out)}",
                type=b.type, text=b.text,
                metadata=b.metadata,
            ))
            cur_page = b.page
            continue
        if b.page != cur_page:
            flush()
            cur_page = b.page
        if not cur_buf:
            cur_buf.append(b.text)
        elif sum(len(x) for x in cur_buf) + len(b.text) + 2 <= max_chars:
            cur_buf.append(b.text)
        else:
            flush()
            cur_page = b.page
            if len(b.text) <= max_chars:
                cur_buf = [b.text]
            else:
                for k in range(0, len(b.text), max_chars):
                    out.append(Block(
                        doc_id=b.doc_id, page=b.page,
                        block_id=f"chunk_{len(out)}",
                        type="text", text=b.text[k:k + max_chars],
                    ))
                cur_buf = []
    flush()
    return out


# ────────────────────────────── ingest write ──────────────────────────────
def write_chunks_to_pg(job_id: str, doc_id: str, file_name: str,
                       chunks: list[Block]) -> tuple[int, int]:
    """Returns (inserted_pg, embedded_count). Idempotent via write_log."""
    from app.agent.tools import _get_embedding_svc
    emb = _get_embedding_svc()
    inserted = 0
    embedded = 0
    conn = _conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """INSERT INTO document_registry
                       (doc_id, file_name, total_pages, status, text_chunk_count, imported_at)
                   VALUES (%s, %s, %s, 'ingesting', %s, NOW())
                   ON CONFLICT (doc_id) DO UPDATE
                     SET text_chunk_count = EXCLUDED.text_chunk_count,
                         status = 'ingesting'""",
                (doc_id, file_name,
                 max((c.page for c in chunks), default=1),
                 len(chunks)),
            )
            conn.commit()

            for i, c in enumerate(chunks):
                key = str(i)
                if write_log_has(job_id, "pg_chunk", key):
                    inserted += 1
                    embedded += 1
                    continue
                try:
                    vec = emb.encode(c.text)
                    vec_list = vec.tolist() if hasattr(vec, "tolist") else list(vec)
                    embedded += 1
                except Exception as ee:
                    logger.warning("embed failed chunk %s: %s", i, ee)
                    vec_list = None

                try:
                    cur.execute(
                        """INSERT INTO text_chunks
                             (doc_id, file_name, chunk_index, content, page_number,
                              section, metadata, embedding, created_at)
                           VALUES (%s, %s, %s, %s, %s, %s, %s::jsonb, %s, NOW())
                           ON CONFLICT (doc_id, file_name, chunk_index) DO UPDATE
                             SET content = EXCLUDED.content,
                                 page_number = EXCLUDED.page_number,
                                 embedding = COALESCE(EXCLUDED.embedding, text_chunks.embedding),
                                 metadata = EXCLUDED.metadata""",
                        (doc_id, file_name, i, c.text, c.page, "",
                         json.dumps({**c.metadata, "job_id": job_id,
                                     "block_type": c.type}, ensure_ascii=False),
                         vec_list),
                    )
                    conn.commit()
                    write_log_put(job_id, "pg_chunk", key)
                    inserted += 1
                except Exception as ce:
                    conn.rollback()
                    logger.warning("pg insert failed chunk %s: %s", i, ce)

            cur.execute(
                "UPDATE document_registry SET status='ready' WHERE doc_id=%s",
                (doc_id,),
            )
            conn.commit()
    finally:
        _release(conn)
    return inserted, embedded


# ────────────────────────────── orchestrator ──────────────────────────────
async def run_ingest_job(job_id: str) -> None:
    started = time.time()
    job = job_get(job_id)
    if not job:
        logger.error("job %s not found", job_id)
        return
    file_path = Path(job["file_path"])
    file_name = job["file_name"]
    suffix = file_path.suffix.lower()
    doc_id = job["doc_id"] or job_id

    job_update(job_id, status="extracting", phase="extract", progress_pct=10,
               started_at=time.strftime("%Y-%m-%d %H:%M:%S"))

    try:
        # 1) extract
        blocks: list[Block] = []
        extractor = ""
        if suffix in (".txt", ".md"):
            blocks, extractor = extract_text_file(file_path, doc_id)
        elif suffix == ".csv":
            blocks, extractor = extract_csv(file_path, doc_id)
        elif suffix in (".xlsx", ".xlsm"):
            blocks, extractor = extract_xlsx(file_path, doc_id)
        elif suffix == ".pdf":
            res = extract_pdf_native(file_path, doc_id, job_id=job_id, file_name=file_name)
            if res:
                blocks, extractor = res
            else:
                # scanned PDF → OCR
                blocks, extractor = await extract_pdf_or_image_via_ocr(
                    file_path, doc_id, file_name, job_id=job_id)
        elif suffix in (".png", ".jpg", ".jpeg"):
            blocks, extractor = await extract_pdf_or_image_via_ocr(
                file_path, doc_id, file_name, job_id=job_id)
        else:
            raise RuntimeError(f"unsupported extension: {suffix}")

        text_chars = sum(len(b.text) for b in blocks)
        job_update(job_id, blocks_total=len(blocks), text_chars=text_chars,
                   extractor=extractor, phase="chunk", progress_pct=40)

        if not blocks:
            raise RuntimeError("extractor produced 0 blocks")

        # 2) chunk
        chunks = chunk_blocks(blocks, max_chars=400)
        job_update(job_id, phase="embed", progress_pct=60)

        # 3) write to PG (with embedding) — idempotent
        inserted, embedded = write_chunks_to_pg(job_id, doc_id, file_name, chunks)
        job_update(job_id, chunks_pg=inserted, vectors_qdrant=embedded,
                   phase="index", progress_pct=80)

        # 4) replicate to Qdrant + Neo4j (idempotent via write_log)
        qdrant_ok = 0
        triples = 0
        neo_drv = _neo4j_session()
        try:
            from app.agent.tools import _get_embedding_svc
            emb = _get_embedding_svc()
            for i, c in enumerate(chunks):
                # Qdrant
                if not write_log_has(job_id, "qdrant", str(i)):
                    try:
                        v = emb.encode(c.text)
                        v_list = v.tolist() if hasattr(v, "tolist") else list(v)
                        if qdrant_upsert(doc_id, file_name, i, c.text, v_list,
                                         c.page, c.type, job_id):
                            write_log_put(job_id, "qdrant", str(i))
                            qdrant_ok += 1
                    except Exception as ee:
                        logger.warning("qdrant chunk %s failed: %s", i, ee)
                else:
                    qdrant_ok += 1
                # Neo4j
                if neo_drv is not None:
                    if not write_log_has(job_id, "neo4j", str(i)):
                        n = neo4j_write_chunk(neo_drv, doc_id, file_name, i,
                                              c.text, c.page, c.type)
                        if n > 0:
                            write_log_put(job_id, "neo4j", str(i))
                            triples += n
                    else:
                        triples += 2  # estimate for already-written chunk
        finally:
            if neo_drv is not None:
                try:
                    neo_drv.close()
                except Exception:
                    pass

        job_update(job_id, vectors_qdrant=qdrant_ok, triples_neo4j=triples,
                   progress_pct=95)

        # 5) finalize
        duration_ms = int((time.time() - started) * 1000)
        job_update(job_id, status="done", phase="done", progress_pct=100,
                   duration_ms=duration_ms,
                   finished_at=time.strftime("%Y-%m-%d %H:%M:%S"))
        logger.info("job %s done: extractor=%s blocks=%d chunks=%d "
                    "qdrant=%d neo4j=%d ms=%d",
                    job_id, extractor, len(blocks), inserted,
                    qdrant_ok, triples, duration_ms)

    except Exception as e:
        duration_ms = int((time.time() - started) * 1000)
        job_update(job_id, status="failed", phase="error",
                   error=str(e)[:1000], duration_ms=duration_ms,
                   finished_at=time.strftime("%Y-%m-%d %H:%M:%S"))
        logger.exception("job %s failed: %s", job_id, e)
